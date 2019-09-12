import openpyxl


def create_query_results_sheets_in_workbook(wb, qry_text):
    """ formats the workbook with sheets named Query and Results
        writes query to sheet

    Args:

        wb (openpyxl workbook)
        qry_text (str): The query that will be executed

    Returns:
        None
    """
    # rename the single sheet to be 'Query'
    sheet = wb['Sheet']
    sheet.title = 'Query'
    # save the query text to cell A1
    sheet['A1'] = qry_text
    # add a new sheet at the first position, and call it 'Results'
    wb.create_sheet(index=0, title='Results')


def get_workbook():
    """ returns a openpyxl workbook,
        so that calling program does not need to import openpyxl

    Returns:
        workbook object
    """
    return openpyxl.Workbook()


def save_workbook_to_file(wb, path, filename):
    """ saves workbook object as XLSX

    Args:
        wb (openpyxl workbook object)
        path (str) - location of saved file
        filename (str) - name of saved file
    Returns:
        None
    """
    # create and write file
    try:
        wb.save(path + '\\' + filename)
    except FileNotFoundError:
        raise Exception(f'File {filename} could not be saved at location {path}.')


def write_results_to_workbook(wb, result_set, results_layout):
    """ execute query and save results into workbook object

    Args:
        wb (openpyxl workbook object)
        result_set (dictionary): results to be written to the workbook
        results_layout (dictionary list): Headers and Fields of what we want in the workbook
    Returns:
        None
    """

    # use results sheet
    sheet = wb['Results']
    # start at row 1
    current_row = 1
    # write header at row 1
    # using Dictionary Header keys, write values to proper columns
    # also validate all column numerics
    for column_dictionary in results_layout:
        column = column_dictionary['column']
        # check for numeric value
        assert str(column).isnumeric(), 'Column "' + str(column) + '" must be a number.'
        # convert to an integer just in case
        column = int(column)
        assert 1 <= column <= 16384, 'Column ' + str(column) + ' must be a number between 1 and 16384.'
        sheet.cell(row=current_row, column=column).value = column_dictionary['header']
    # move to next row
    current_row += 1

    # write data starting at row 2
    for row in result_set:
        # using Dictionary Field keys, write values to proper columns
        for column_dictionary in results_layout:
            column = int(column_dictionary['column'])
            assert column_dictionary['field'] in row, \
                'Field "' + column_dictionary['field'] + '" was not found in the query results.'
            sheet.cell(row=current_row, column=column).value = row[column_dictionary['field']]
        # move to next row fro next query result
        current_row += 1
